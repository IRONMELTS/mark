from openpyxl import load_workbook
import matplotlib.pyplot as plt

wb=load_workbook('marklist.xlsx')
ws=wb.active
h=ws[1]
a=''
sub=False

print()
print("NOTE THAT THE SUBJECT NAME IS CASE SENSITIVE")
print("NOTE THAT THE SUBJECT NAME IS ENTERED AS THE WAY ENTERED BEFORE")
print()

while not a.lower()=='stop':
    a=input('Enter the name of subject to get the chart:')
    for j in h:
        if j.value == a.lower():
            print(j.column_letter)
            plt.figure(figsize=(6,5))
            eng=ws[f'{j.column_letter}']
            k=[]
            for i in eng:
                if not isinstance(i.value, str):
                    k.append(i.value)

            acima90=0
            acima80=0
            acima60=0
            abaixo60=0
            l=['Above 90%','Above 80%','Above 60%','Belove 60%']
            for cell in k:
                if cell > 90:
                    acima90+=1
                elif cell <=90 and cell > 80:
                    acima80+=1
                elif cell <=80 and cell >60:
                    acima60+=1
                else:
                    abaixo60+=1
                    
            val=[acima90,acima80,acima60,abaixo60]

            print(val)
            plt.pie(val,labels=l,autopct='%1.1f%%')
            plt.title(a)
            plt.show()
            sub=True
            break
    
    if sub==False:
        print("Enter valid subject")
        print("NOTE THAT THE SUBJECT NAME IS CASE SENSITIVE")
        print("NOTE THAT THE SUBJECT NAME IS ENTERED AS THE WAY ENTERED BEFORE")
